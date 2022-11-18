import numpy as np
from openpyxl import Workbook, load_workbook

#---------------------------------------------------
wb = load_workbook('aaa.xlsx')
ws = wb.active

#in ra giá trị trong 1 cột
ws = wb['Sheet1']
col = ws['A']
col_value = []
for i in col:
    col_value.append(i.value)
# print (col_value)

#in ra giá trị trong 1 hàng
ws = wb['Sheet1']
row = ws[1]
row_value = []
for i in row:
    row_value.append(i.value)
print ('giá trị hàng:',row_value)
for demhang in range(0, len(col)):
    # print (demhang)
    None


#so sánh 1 phần tử với hàng vừa in ra
def timdulieu(name, row_value):
    result = -1
    for j in range (0, len(row_value)):
            if row_value[j]==name:
                result = j
                break
    return result

def check_user_input(input):
    try:
        # Convert it into integer
        int(input)
        return 1
    except ValueError:
        return -1



def nhapten():
        wb1 = load_workbook(filename = 'bbb.xlsx')
        sheet_ranges = wb1['Sheet']
        row = sheet_ranges[1]

        row_value1 = []
        wb2 = Workbook()
        dest_filename = 'bbb.xlsx'
        ws2 = wb2.active
        ws2.title = "Sheet"
        for i in row:
            row_value1.append(i.value)
            #  print (row_value1)
        for m in row_value1:
            None
            # print (m)
            
    # while True :
            # print("Nhập tên muốn tìm:")
        name = m
        if (check_user_input(name) ==1 ):
            name = int(name)
        else:
            None
        # print (type(name))
        index = timdulieu(name, row_value)
                
        if (index == -1):
            print("Không tìm thấy")
        else:
            print("Tìm thấy tại vị trí ", index+1)
            # print(index+1)
        for colu in ws.iter_rows(min_row=1, max_row = demhang+1, min_col=1, max_col=index+1,  values_only=True):
            print(colu)
        # break
        
            # for row in range(1, demhang+2):
            #     print (row)
            # for col in range(1, index+1):
            ws2.append(colu)

        wb2.save(filename = dest_filename)
nhapten()
